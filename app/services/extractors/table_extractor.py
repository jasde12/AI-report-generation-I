from __future__ import annotations

from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.models.schemas import (
    ExtractedSource,
    ExtractedTable,
    MergedCellRange,
    SourceBlock,
    TableCell,
)


class TableExtractor:
    HEADER_SCAN_LIMIT = 3
    CSV_ENCODINGS = ("utf-8-sig", "utf-8", "cp950")

    def extract(self, file_path: Path) -> ExtractedSource:
        suffix = file_path.suffix.lower()
        if suffix == ".csv":
            tables = [self._extract_csv(file_path)]
            metadata = {"sheet_names": ["csv"]}
        elif suffix == ".xlsx":
            tables, sheet_names = self._extract_xlsx(file_path)
            metadata = {"sheet_names": sheet_names}
        else:
            raise ValueError(f"不支援的表格格式: {file_path.name}")

        blocks: list[SourceBlock] = []
        warnings: list[str] = []
        for index, table in enumerate(tables, start=1):
            table_title = table.sheet_name or table.title or f"Table {index}"
            blocks.append(
                SourceBlock(
                    order=index,
                    block_type="table",
                    table_index=index,
                    table_title=table_title,
                    table_markdown=table.markdown,
                )
            )
            warnings.extend(
                f"{table_title}: {warning}"
                for warning in table.warnings
                if any(keyword in warning for keyword in ("未能", "無法"))
            )

        return ExtractedSource(
            type="table",
            source_name=file_path.name,
            tables=tables,
            blocks=blocks,
            warnings=warnings,
            metadata=metadata,
        )

    def _extract_csv(self, file_path: Path) -> ExtractedTable:
        last_error: Exception | None = None
        for encoding in self.CSV_ENCODINGS:
            try:
                dataframe = pd.read_csv(
                    file_path,
                    dtype=object,
                    encoding=encoding,
                    header=None,
                )
                return self._dataframe_to_table(dataframe, title=file_path.stem)
            except Exception as exc:
                last_error = exc

        raise ValueError(f"CSV 讀取失敗: {file_path.name}") from last_error

    def _extract_xlsx(self, file_path: Path) -> tuple[list[ExtractedTable], list[str]]:
        try:
            workbook = pd.ExcelFile(file_path, engine="openpyxl")
            openpyxl_workbook = load_workbook(file_path, data_only=True)
        except Exception as exc:
            raise ValueError(f"Excel 讀取失敗: {file_path.name}") from exc

        sheet_names = list(workbook.sheet_names)
        try:
            tables: list[ExtractedTable] = []
            for sheet_name in sheet_names:
                dataframe = workbook.parse(sheet_name=sheet_name, dtype=object, header=None)
                worksheet = openpyxl_workbook[sheet_name]
                tables.append(
                    self._dataframe_to_table(
                        dataframe=dataframe,
                        title=sheet_name,
                        sheet_name=sheet_name,
                        worksheet=worksheet,
                    )
                )
        finally:
            workbook.close()
            openpyxl_workbook.close()

        return tables, sheet_names

    def _dataframe_to_table(
        self,
        dataframe: pd.DataFrame,
        title: str,
        sheet_name: str | None = None,
        worksheet: Worksheet | None = None,
    ) -> ExtractedTable:
        raw_matrix, row_offset, column_offset = self._dataframe_to_matrix(dataframe)
        if not raw_matrix:
            return ExtractedTable(
                title=title,
                sheet_name=sheet_name,
                columns=[],
                rows=[],
                raw_rows=[],
                header_rows=[],
                markdown="(empty table)",
                row_count=0,
                column_count=0,
                cells=[],
                merged_ranges=[],
                header_depth=0,
                warnings=[],
            )

        merged_ranges = self._extract_merged_ranges(
            worksheet=worksheet,
            row_offset=row_offset,
            column_offset=column_offset,
            row_count=len(raw_matrix),
            column_count=max(len(row) for row in raw_matrix),
        )
        display_matrix = self._apply_merged_values(raw_matrix, merged_ranges)

        header_depth = self._infer_header_depth(display_matrix, merged_ranges)
        columns, normalized_rows, warnings = self._normalize_table(display_matrix, header_depth)

        markdown = self._build_markdown(
            columns=columns,
            normalized_rows=normalized_rows,
            display_matrix=display_matrix,
        )

        rows: list[Any]
        if columns:
            rows = normalized_rows
        else:
            rows = display_matrix

        return ExtractedTable(
            title=title,
            sheet_name=sheet_name,
            columns=columns,
            rows=rows,
            raw_rows=raw_matrix,
            header_rows=self._matrix_to_string_rows(display_matrix[:header_depth]),
            markdown=markdown,
            row_count=len(raw_matrix),
            column_count=max(len(row) for row in raw_matrix),
            cells=self._build_cells(display_matrix),
            merged_ranges=merged_ranges,
            header_depth=header_depth,
            warnings=warnings,
        )

    def _dataframe_to_matrix(
        self,
        dataframe: pd.DataFrame,
    ) -> tuple[list[list[Any]], int, int]:
        normalized = dataframe.where(pd.notna(dataframe), None)
        rows = [
            [self._to_native(value) for value in row]
            for row in normalized.values.tolist()
        ]
        return self._trim_outer_empty(rows)

    def _trim_outer_empty(
        self,
        rows: list[list[Any]],
    ) -> tuple[list[list[Any]], int, int]:
        if not rows:
            return [], 0, 0

        row_indexes = [
            index
            for index, row in enumerate(rows)
            if any(not self._is_empty(cell) for cell in row)
        ]
        if not row_indexes:
            return [], 0, 0

        top = row_indexes[0]
        bottom = row_indexes[-1]
        sliced_rows = rows[top : bottom + 1]

        width = max(len(row) for row in sliced_rows)
        padded_rows = [row + [None] * (width - len(row)) for row in sliced_rows]

        column_indexes = [
            index
            for index in range(width)
            if any(not self._is_empty(row[index]) for row in padded_rows)
        ]
        if not column_indexes:
            return [], top, 0

        left = column_indexes[0]
        right = column_indexes[-1]
        trimmed_rows = [row[left : right + 1] for row in padded_rows]
        return trimmed_rows, top, left

    def _extract_merged_ranges(
        self,
        worksheet: Worksheet | None,
        row_offset: int,
        column_offset: int,
        row_count: int,
        column_count: int,
    ) -> list[MergedCellRange]:
        if worksheet is None:
            return []

        sheet_top = row_offset + 1
        sheet_bottom = row_offset + row_count
        sheet_left = column_offset + 1
        sheet_right = column_offset + column_count

        merged_ranges: list[MergedCellRange] = []
        for merged_range in worksheet.merged_cells.ranges:
            if (
                merged_range.max_row < sheet_top
                or merged_range.min_row > sheet_bottom
                or merged_range.max_col < sheet_left
                or merged_range.min_col > sheet_right
            ):
                continue

            start_row = max(merged_range.min_row, sheet_top) - row_offset
            end_row = min(merged_range.max_row, sheet_bottom) - row_offset
            start_column = max(merged_range.min_col, sheet_left) - column_offset
            end_column = min(merged_range.max_col, sheet_right) - column_offset

            merged_ranges.append(
                MergedCellRange(
                    start_row=start_row,
                    start_column=start_column,
                    end_row=end_row,
                    end_column=end_column,
                )
            )

        return merged_ranges

    def _apply_merged_values(
        self,
        rows: list[list[Any]],
        merged_ranges: list[MergedCellRange],
    ) -> list[list[Any]]:
        display_rows = [row[:] for row in rows]
        for merged_range in merged_ranges:
            top_left_value = display_rows[merged_range.start_row - 1][merged_range.start_column - 1]
            if self._is_empty(top_left_value):
                continue

            for row_index in range(merged_range.start_row - 1, merged_range.end_row):
                for column_index in range(merged_range.start_column - 1, merged_range.end_column):
                    if self._is_empty(display_rows[row_index][column_index]):
                        display_rows[row_index][column_index] = top_left_value

        return display_rows

    def _infer_header_depth(
        self,
        rows: list[list[Any]],
        merged_ranges: list[MergedCellRange],
    ) -> int:
        max_depth = min(self.HEADER_SCAN_LIMIT, len(rows) - 1)
        if max_depth <= 0:
            return 0

        best_depth = 0
        best_score = 0.0
        merged_header_bonus = 0.35 if any(
            merged_range.start_row <= self.HEADER_SCAN_LIMIT
            and merged_range.end_row > merged_range.start_row
            for merged_range in merged_ranges
        ) else 0.0

        for depth in range(1, max_depth + 1):
            header_rows = rows[:depth]
            next_row = rows[depth]
            flattened_headers = self._combine_header_rows(header_rows)
            normalized_headers = self._normalize_columns(flattened_headers)

            informative_ratio = sum(
                1 for header in normalized_headers if not header.startswith("column_")
            ) / max(len(normalized_headers), 1)
            duplicate_penalty = len(normalized_headers) - len(set(normalized_headers))
            header_score = sum(self._score_header_row(row) for row in header_rows) / depth
            data_score = self._score_data_row(next_row)

            score = (
                header_score * 0.9
                + data_score * 0.8
                + informative_ratio * 0.8
                + (merged_header_bonus if depth > 1 else 0.0)
                - duplicate_penalty * 0.08
                - (depth - 1) * 0.05
            )

            if score > best_score:
                best_score = score
                best_depth = depth

        if best_score < 1.15 or self._score_header_row(rows[0]) < 0.45:
            return 0

        return best_depth

    def _normalize_table(
        self,
        rows: list[list[Any]],
        header_depth: int,
    ) -> tuple[list[str], list[dict[str, Any]], list[str]]:
        if header_depth <= 0:
            return [], [], ["未能安全辨識表頭，已保留原始表格結構。"]

        header_rows = rows[:header_depth]
        data_rows = rows[header_depth:]
        columns = self._normalize_columns(self._combine_header_rows(header_rows))

        normalized_rows: list[dict[str, Any]] = []
        for row in data_rows:
            if self._is_empty_row(row):
                continue

            padded_row = row + [None] * (len(columns) - len(row))
            normalized_rows.append(
                {
                    column: self._to_native(padded_row[index])
                    for index, column in enumerate(columns)
                }
            )

        warnings: list[str] = []
        if header_depth > 1:
            warnings.append(f"偵測到 {header_depth} 層表頭，已展平成單層欄位。")
        if any(column.startswith("column_") for column in columns):
            warnings.append("部分欄位名稱無法安全辨識，已使用通用欄位名稱。")

        return columns, normalized_rows, warnings

    def _combine_header_rows(self, rows: list[list[Any]]) -> list[str]:
        if not rows:
            return []

        width = max(len(row) for row in rows)
        combined_headers: list[str] = []
        for column_index in range(width):
            parts: list[str] = []
            for row in rows:
                value = row[column_index] if column_index < len(row) else None
                text = self._stringify(value)
                if not text:
                    continue
                if not parts or parts[-1] != text:
                    parts.append(text)

            combined_headers.append(" / ".join(parts) if parts else f"column_{column_index + 1}")

        return combined_headers

    def _build_markdown(
        self,
        columns: list[str],
        normalized_rows: list[dict[str, Any]],
        display_matrix: list[list[Any]],
    ) -> str:
        if columns:
            return self._records_to_markdown(columns, normalized_rows)
        return self._matrix_to_markdown(display_matrix)

    def _build_cells(self, rows: list[list[Any]]) -> list[TableCell]:
        cells: list[TableCell] = []
        for row_index, row in enumerate(rows, start=1):
            for column_index, value in enumerate(row, start=1):
                if self._is_empty(value):
                    continue
                cells.append(
                    TableCell(
                        row_index=row_index,
                        column_index=column_index,
                        text=self._stringify(value),
                    )
                )
        return cells

    def _matrix_to_markdown(self, rows: list[list[Any]]) -> str:
        if not rows:
            return "(empty table)"

        width = max(len(row) for row in rows)
        headers = [f"column_{index + 1}" for index in range(width)]
        markdown_lines = [
            "| " + " | ".join(headers) + " |",
            "| " + " | ".join(["---"] * width) + " |",
        ]
        for row in rows:
            padded_row = row + [None] * (width - len(row))
            markdown_lines.append(
                "| "
                + " | ".join(self._stringify(value) for value in padded_row)
                + " |"
            )
        return "\n".join(markdown_lines)

    def _records_to_markdown(
        self,
        columns: list[str],
        rows: list[dict[str, Any]],
    ) -> str:
        if not columns:
            return "(empty table)"

        markdown_lines = [
            "| " + " | ".join(columns) + " |",
            "| " + " | ".join(["---"] * len(columns)) + " |",
        ]
        for row in rows:
            markdown_lines.append(
                "| "
                + " | ".join(self._stringify(row.get(column)) for column in columns)
                + " |"
            )
        return "\n".join(markdown_lines)

    def _matrix_to_string_rows(self, rows: list[list[Any]]) -> list[list[str]]:
        return [[self._stringify(value) for value in row] for row in rows]

    def _score_header_row(self, row: list[Any]) -> float:
        non_empty_values = [value for value in row if not self._is_empty(value)]
        if not non_empty_values:
            return 0.0

        text_like = sum(1 for value in non_empty_values if not self._is_number_like(value))
        short_labels = sum(1 for value in non_empty_values if len(self._stringify(value)) <= 20)
        return (text_like / len(non_empty_values)) * 0.7 + (
            short_labels / len(non_empty_values)
        ) * 0.3

    def _score_data_row(self, row: list[Any]) -> float:
        non_empty_values = [value for value in row if not self._is_empty(value)]
        if not non_empty_values:
            return 0.0

        numeric_like = sum(1 for value in non_empty_values if self._is_number_like(value))
        marker_like = sum(1 for value in non_empty_values if self._is_marker_value(value))
        long_text = sum(1 for value in non_empty_values if len(self._stringify(value)) > 12)
        first_value = self._first_non_empty_value(row)
        label_bonus = 0.15 if first_value is not None and not self._is_number_like(first_value) else 0.0

        return min(
            1.0,
            ((numeric_like + marker_like) / len(non_empty_values))
            + (long_text / len(non_empty_values)) * 0.35
            + label_bonus,
        )

    def _normalize_columns(self, columns: list[str]) -> list[str]:
        normalized: list[str] = []
        seen: dict[str, int] = {}
        for index, column in enumerate(columns):
            base_name = column.strip() if column is not None else f"column_{index + 1}"
            if not base_name or base_name.lower().startswith("unnamed:"):
                base_name = f"column_{index + 1}"

            current_count = seen.get(base_name, 0)
            seen[base_name] = current_count + 1
            if current_count:
                normalized.append(f"{base_name}_{current_count + 1}")
            else:
                normalized.append(base_name)
        return normalized

    def _is_empty_row(self, row: list[Any]) -> bool:
        return all(self._is_empty(value) for value in row)

    def _is_empty(self, value: Any) -> bool:
        if value is None:
            return True
        if isinstance(value, str):
            return not value.strip()
        return False

    def _is_number_like(self, value: Any) -> bool:
        if isinstance(value, bool):
            return False
        if isinstance(value, (int, float)) and value is not None:
            return True
        if not isinstance(value, str):
            return False

        candidate = value.strip().replace(",", "").replace("%", "")
        if not candidate:
            return False
        try:
            float(candidate)
        except ValueError:
            return False
        return True

    def _is_marker_value(self, value: Any) -> bool:
        if value is None:
            return False
        text = self._stringify(value).lower()
        return text in {"v", "y", "yes", "x", "ok", "true", "false", "■", "□", "○"}

    def _first_non_empty_value(self, row: list[Any]) -> Any | None:
        for value in row:
            if not self._is_empty(value):
                return value
        return None

    def _to_native(self, value: Any) -> Any:
        if hasattr(value, "item"):
            try:
                return value.item()
            except Exception:
                return value
        return value

    def _stringify(self, value: Any) -> str:
        if value is None:
            return ""
        return str(value).replace("\n", " ").strip()
